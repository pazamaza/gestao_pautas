from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from alunos.models import Aluno
from pautas.models import Nota


CABECALHOS = [
    'numero_processo',
    'aluno',
    'mac',
    'npp',
    'npt',
    'observacao',
]


def situacao_por_media(media):
    if media >= 10:
        return 'Aprovado'
    if media >= 8:
        return 'Exame'
    return 'Reprovado'


def criar_modelo_excel(avaliacao):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Pauta'

    sheet['A1'] = 'Gestao de Pautas'
    sheet['A2'] = str(avaliacao.atribuicao)
    sheet['A3'] = str(avaliacao.periodo)
    sheet['A1'].font = Font(bold=True, size=14)

    sheet.append([])
    sheet.append(CABECALHOS)

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in sheet[5]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    alunos = (
        Aluno.objects
        .filter(turma=avaliacao.atribuicao.turma, ativo=True)
        .order_by('nome')
    )

    for aluno in alunos:
        nota = Nota.objects.filter(
            avaliacao=avaliacao,
            aluno=aluno,
        ).first()

        sheet.append([
            aluno.numero_processo,
            aluno.nome,
            nota.mac if nota else '',
            nota.npp if nota else '',
            nota.npt if nota else '',
            nota.observacao if nota else '',
        ])

    for column in range(1, len(CABECALHOS) + 1):
        letter = get_column_letter(column)
        sheet.column_dimensions[letter].width = 18
    sheet.column_dimensions['B'].width = 34
    sheet.column_dimensions['F'].width = 40

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def exportar_pauta_excel(avaliacao, notas):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Pauta Trimestral'

    sheet.append(['Gestao de Pautas'])
    sheet.append([str(avaliacao.atribuicao)])
    sheet.append([str(avaliacao.periodo)])
    sheet.append([])
    sheet.append([
        'Nº',
        'Nº Processo',
        'Aluno',
        'MAC',
        'NPP',
        'NPT',
        'MT',
        'Situacao',
        'Observacao',
    ])

    for cell in sheet[5]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='D9EAF7')
        cell.alignment = Alignment(horizontal='center')

    for indice, nota in enumerate(notas, start=1):
        sheet.append([
            indice,
            nota.aluno.numero_processo,
            nota.aluno.nome,
            nota.mac,
            nota.npp,
            nota.npt,
            nota.mt,
            situacao_por_media(nota.mt),
            nota.observacao or '',
        ])

    larguras = [8, 18, 34, 10, 10, 10, 10, 16, 36]
    for index, largura in enumerate(larguras, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = largura

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def importar_notas_excel(avaliacao, arquivo):
    workbook = load_workbook(filename=arquivo, data_only=True)
    sheet = workbook.active
    cabecalhos = [
        str(cell.value).strip().lower() if cell.value else ''
        for cell in sheet[5]
    ]

    indices = {nome: cabecalhos.index(nome) for nome in CABECALHOS if nome in cabecalhos}
    obrigatorios = {'numero_processo', 'mac', 'npp', 'npt'}
    faltando = obrigatorios - set(indices)
    if faltando:
        return {
            'criados': 0,
            'atualizados': 0,
            'erros': [f"Colunas em falta: {', '.join(sorted(faltando))}."],
        }

    criados = 0
    atualizados = 0
    erros = []

    with transaction.atomic():
        for numero_linha, row in enumerate(sheet.iter_rows(min_row=6, values_only=True), start=6):
            if not any(row):
                continue

            numero_processo = str(row[indices['numero_processo']] or '').strip()
            if not numero_processo:
                erros.append(f'Linha {numero_linha}: numero_processo vazio.')
                continue

            try:
                aluno = Aluno.objects.get(
                    numero_processo=numero_processo,
                    turma=avaliacao.atribuicao.turma,
                )
            except Aluno.DoesNotExist:
                erros.append(
                    f'Linha {numero_linha}: aluno {numero_processo} nao encontrado nesta turma.'
                )
                continue

            try:
                mac = converter_nota(row[indices['mac']])
                npp = converter_nota(row[indices['npp']])
                npt = converter_nota(row[indices['npt']])
            except ValueError as exc:
                erros.append(f'Linha {numero_linha}: {exc}')
                continue

            observacao = ''
            if 'observacao' in indices and row[indices['observacao']] is not None:
                observacao = str(row[indices['observacao']]).strip()

            _, criado = Nota.objects.update_or_create(
                avaliacao=avaliacao,
                aluno=aluno,
                defaults={
                    'mac': mac,
                    'npp': npp,
                    'npt': npt,
                    'observacao': observacao,
                },
            )

            if criado:
                criados += 1
            else:
                atualizados += 1

    return {
        'criados': criados,
        'atualizados': atualizados,
        'erros': erros,
    }


def exportar_pauta_final_excel(turma, ano_letivo, disciplinas, linhas):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Pauta Final'

    sheet['A1'] = 'Gestao de Pautas'
    sheet['A1'].font = Font(bold=True, size=14)
    sheet['A2'] = f'Pauta Final - {turma}'
    sheet['A3'] = (
        f'Sala: {turma.sala or "-"}  |  '
        f'Periodo: {turma.get_periodo_display() or "-"}  |  '
        f'Ano Letivo: {ano_letivo}'
    )

    linha_disciplinas = 5
    linha_subcabecalhos = 6

    sheet.cell(row=linha_disciplinas, column=1, value='Nº')
    sheet.cell(row=linha_disciplinas, column=2, value='Aluno')
    sheet.merge_cells(
        start_row=linha_disciplinas, start_column=1, end_row=linha_subcabecalhos, end_column=1
    )
    sheet.merge_cells(
        start_row=linha_disciplinas, start_column=2, end_row=linha_subcabecalhos, end_column=2
    )

    coluna = 3
    for disciplina in disciplinas:
        sheet.cell(row=linha_disciplinas, column=coluna, value=str(disciplina))
        sheet.merge_cells(
            start_row=linha_disciplinas, start_column=coluna,
            end_row=linha_disciplinas, end_column=coluna + 3,
        )
        for indice, rotulo in enumerate(['MFD', 'NE', 'MF', 'NER']):
            sheet.cell(row=linha_subcabecalhos, column=coluna + indice, value=rotulo)
        coluna += 4

    sheet.cell(row=linha_disciplinas, column=coluna, value='Situação Geral')
    sheet.merge_cells(
        start_row=linha_disciplinas, start_column=coluna,
        end_row=linha_subcabecalhos, end_column=coluna,
    )
    total_colunas = coluna

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for linha_cabecalho in sheet.iter_rows(
        min_row=linha_disciplinas, max_row=linha_subcabecalhos, min_col=1, max_col=total_colunas
    ):
        for cell in linha_cabecalho:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for indice, linha in enumerate(linhas, start=1):
        valores = [indice, str(linha['aluno'])]
        for resultado in linha['celulas']:
            if resultado:
                valores.extend([
                    resultado.mf,
                    resultado.exame if resultado.exame is not None else '',
                    resultado.nota_final if resultado.nota_final is not None else '',
                    resultado.nota_recurso if resultado.nota_recurso is not None else '',
                ])
            else:
                valores.extend(['', '', '', ''])
        situacao_anual = linha['situacao_anual']
        valores.append(situacao_anual.situacao if situacao_anual else '')
        sheet.append(valores)

    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 30
    for coluna_indice in range(3, total_colunas):
        sheet.column_dimensions[get_column_letter(coluna_indice)].width = 8
    sheet.column_dimensions[get_column_letter(total_colunas)].width = 24

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def converter_nota(valor):
    try:
        nota = Decimal(str(valor).replace(',', '.'))
    except (InvalidOperation, AttributeError):
        raise ValueError(f'nota invalida: {valor}')

    if nota < 0 or nota > 20:
        raise ValueError(f'nota fora do intervalo 0-20: {valor}')

    return nota.quantize(Decimal('0.1'))
